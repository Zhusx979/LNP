from __future__ import annotations

import csv
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
import re
from typing import Iterable

from openpyxl import Workbook, load_workbook


EXPECTED_COUNTS = {"A": 10, "B": 10, "C": 20, "D": 20}
TOTAL_ROWS = 10 * 10 * 20 * 20
INPUT_FILENAME = "gradient for waiting hub.xlsx"
OUTPUT_FILENAME = "候选库4万.xlsx"


@dataclass(frozen=True)
class Component:
    code: str
    smiles: str


@dataclass(frozen=True)
class WaitingHubLibrary:
    a_values: list[Component]
    b_values: list[Component]
    c_values: list[Component]
    d_values: list[Component]


@lru_cache(maxsize=1)
def _rdkit_modules():
    from rdkit import Chem
    from rdkit.Chem import rdchem

    return Chem, rdchem


def rdkit_available() -> bool:
    try:
        _rdkit_modules()
        return True
    except ModuleNotFoundError:
        return False


def canonicalize_smiles(smiles: str) -> str:
    Chem, _ = _rdkit_modules()
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError(f"Invalid SMILES: {smiles}")
    return Chem.MolToSmiles(mol)


def _find_isocyanide_atoms(mol) -> tuple[int, int]:
    Chem, _ = _rdkit_modules()
    match = mol.GetSubstructMatches(Chem.MolFromSmarts("[N+]#[C-]"))
    if not match:
        raise ValueError("Isocyanide group not found")
    return match[0]


def _find_carbonyl_atoms(mol) -> tuple[int, int]:
    _, rdchem = _rdkit_modules()
    for atom in mol.GetAtoms():
        if atom.GetAtomicNum() != 6:
            continue
        for bond in atom.GetBonds():
            other = bond.GetOtherAtom(atom)
            if (
                bond.GetBondType() == rdchem.BondType.DOUBLE
                and other.GetAtomicNum() == 8
            ):
                return atom.GetIdx(), other.GetIdx()
    raise ValueError("Carbonyl group not found")


def _find_primary_amine_n(mol) -> int:
    for atom in mol.GetAtoms():
        if atom.GetAtomicNum() == 7 and atom.GetTotalNumHs() >= 2:
            return atom.GetIdx()
    raise ValueError("Primary amine nitrogen not found")


def _find_carboxylic_acid_atoms(mol) -> tuple[int, int]:
    Chem, _ = _rdkit_modules()
    matches = mol.GetSubstructMatches(Chem.MolFromSmarts("[CX3](=O)[OX2H1]"))
    if not matches:
        raise ValueError("Carboxylic acid group not found")
    carbonyl_carbon, _, hydroxy_oxygen = matches[0]
    return carbonyl_carbon, hydroxy_oxygen


def _shift_after_removals(index: int, removed_indices: list[int]) -> int:
    return index - sum(1 for removed in removed_indices if removed < index)


def generate_ugi_product_smiles(
    isocyanide_smiles: str,
    carbonyl_smiles: str,
    amine_smiles: str,
    acid_smiles: str,
) -> str:
    Chem, rdchem = _rdkit_modules()

    isocyanide = Chem.MolFromSmiles(isocyanide_smiles)
    carbonyl = Chem.MolFromSmiles(carbonyl_smiles)
    amine = Chem.MolFromSmiles(amine_smiles)
    acid = Chem.MolFromSmiles(acid_smiles)

    if None in (isocyanide, carbonyl, amine, acid):
        raise ValueError("Failed to parse one or more reactant SMILES")

    iso_n, iso_c = _find_isocyanide_atoms(isocyanide)
    carbonyl_c, carbonyl_o = _find_carbonyl_atoms(carbonyl)
    amine_n = _find_primary_amine_n(amine)
    acid_c, acid_oh_o = _find_carboxylic_acid_atoms(acid)

    combo = Chem.CombineMols(
        Chem.CombineMols(Chem.CombineMols(isocyanide, carbonyl), amine),
        acid,
    )
    offsets = [
        0,
        isocyanide.GetNumAtoms(),
        isocyanide.GetNumAtoms() + carbonyl.GetNumAtoms(),
        isocyanide.GetNumAtoms() + carbonyl.GetNumAtoms() + amine.GetNumAtoms(),
    ]

    iso_n += offsets[0]
    iso_c += offsets[0]
    carbonyl_c += offsets[1]
    carbonyl_o += offsets[1]
    amine_n += offsets[2]
    acid_c += offsets[3]
    acid_oh_o += offsets[3]

    rw = Chem.RWMol(combo)
    removed = sorted([carbonyl_o, acid_oh_o], reverse=True)
    for atom_idx in removed:
        rw.RemoveAtom(atom_idx)

    removed_low_to_high = sorted(removed)
    iso_n = _shift_after_removals(iso_n, removed_low_to_high)
    iso_c = _shift_after_removals(iso_c, removed_low_to_high)
    carbonyl_c = _shift_after_removals(carbonyl_c, removed_low_to_high)
    amine_n = _shift_after_removals(amine_n, removed_low_to_high)
    acid_c = _shift_after_removals(acid_c, removed_low_to_high)

    iso_bond = rw.GetBondBetweenAtoms(iso_n, iso_c)
    iso_bond.SetBondType(rdchem.BondType.SINGLE)

    iso_n_atom = rw.GetAtomWithIdx(iso_n)
    iso_n_atom.SetFormalCharge(0)
    iso_n_atom.SetNumExplicitHs(1)
    iso_n_atom.SetNoImplicit(True)
    rw.GetAtomWithIdx(iso_c).SetFormalCharge(0)

    oxygen_idx = rw.AddAtom(Chem.Atom(8))
    rw.AddBond(iso_c, oxygen_idx, rdchem.BondType.DOUBLE)
    rw.AddBond(iso_c, carbonyl_c, rdchem.BondType.SINGLE)
    rw.AddBond(amine_n, carbonyl_c, rdchem.BondType.SINGLE)
    rw.AddBond(amine_n, acid_c, rdchem.BondType.SINGLE)

    product = rw.GetMol()
    Chem.SanitizeMol(product)
    return canonicalize_smiles(Chem.MolToSmiles(product))


def _read_group(sheet, smiles_col: int, code_col: int) -> list[Component]:
    values: list[Component] = []
    for row in range(2, sheet.max_row + 1):
        smiles = sheet.cell(row, smiles_col).value
        code = sheet.cell(row, code_col).value

        if smiles is None and code is None:
            continue
        if smiles is None or code is None:
            raise ValueError(
                f"Incomplete component row {row} for columns {smiles_col}/{code_col}: "
                f"smiles={smiles!r}, code={code!r}"
            )

        values.append(Component(code=str(code).strip(), smiles=str(smiles).strip()))
    return values


def load_waiting_hub_library(workbook_path: Path) -> WaitingHubLibrary:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    a_values = _read_group(sheet, smiles_col=2, code_col=3)
    b_values = _read_group(sheet, smiles_col=4, code_col=5)
    c_values = _read_group(sheet, smiles_col=6, code_col=7)
    d_values = _read_group(sheet, smiles_col=8, code_col=10)

    counts = {
        "A": len(a_values),
        "B": len(b_values),
        "C": len(c_values),
        "D": len(d_values),
    }
    if counts != EXPECTED_COUNTS:
        raise ValueError(f"Unexpected waiting hub counts: {counts}, expected {EXPECTED_COUNTS}")

    return WaitingHubLibrary(
        a_values=a_values,
        b_values=b_values,
        c_values=c_values,
        d_values=d_values,
    )


def _parse_code_number(code: str, prefix: str) -> int:
    match = re.fullmatch(rf"{prefix}(\d+)", code)
    if not match:
        raise ValueError(f"Unexpected component code: {code}")
    return int(match.group(1))


def build_existing_smiles_lookup(
    library: WaitingHubLibrary,
    data_dir: Path,
) -> dict[str, str]:
    generated_paths = {
        "block1": data_dir / "LNP_virtual_lipid_library_generated2.csv",
        "block2": data_dir / "LNP_virtual_lipid_library_generated1.csv",
    }
    for path in generated_paths.values():
        if not path.exists():
            raise FileNotFoundError(f"Fallback library file not found: {path}")

    required_by_source: dict[str, set[str]] = {"block1": set(), "block2": set()}
    original_to_lookup: dict[str, tuple[str, str]] = {}

    for a_item in library.a_values:
        a_num = _parse_code_number(a_item.code, "A")
        source_name = "block1" if a_num <= 20 else "block2"
        remapped_a = a_num if a_num <= 20 else a_num - 20

        for b_item in library.b_values:
            for c_item in library.c_values:
                for d_item in library.d_values:
                    original_combo = f"{a_item.code}{b_item.code}{c_item.code}{d_item.code}"
                    lookup_combo = f"A{remapped_a}{b_item.code}{c_item.code}{d_item.code}"
                    required_by_source[source_name].add(lookup_combo)
                    original_to_lookup[original_combo] = (source_name, lookup_combo)

    found_by_source: dict[tuple[str, str], str] = {}
    for source_name, csv_path in generated_paths.items():
        expected = required_by_source[source_name]
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                combo = row["Combo"]
                if combo in expected:
                    found_by_source[(source_name, combo)] = row["SMILES"]
                    if len([key for key in found_by_source if key[0] == source_name]) == len(expected):
                        break

    if len(found_by_source) != sum(len(values) for values in required_by_source.values()):
        missing = []
        for source_name, combos in required_by_source.items():
            for combo in combos:
                if (source_name, combo) not in found_by_source:
                    missing.append(f"{source_name}:{combo}")
                    if len(missing) >= 5:
                        break
            if missing:
                break
        raise ValueError(f"Missing fallback SMILES for combos like: {missing}")

    return {
        original_combo: found_by_source[lookup_key]
        for original_combo, lookup_key in original_to_lookup.items()
    }


def iter_waiting_hub_rows(
    library: WaitingHubLibrary,
    smiles_lookup: dict[str, str] | None = None,
) -> Iterable[tuple[str, str]]:
    for a_item in library.a_values:
        for b_item in library.b_values:
            for c_item in library.c_values:
                for d_item in library.d_values:
                    combo = f"{a_item.code}{b_item.code}{c_item.code}{d_item.code}"
                    if smiles_lookup is None:
                        smiles = generate_ugi_product_smiles(
                            isocyanide_smiles=a_item.smiles,
                            carbonyl_smiles=b_item.smiles,
                            amine_smiles=c_item.smiles,
                            acid_smiles=d_item.smiles,
                        )
                    else:
                        smiles = smiles_lookup[combo]
                    yield combo, smiles


def export_waiting_hub_excel(
    output_path: Path,
    library: WaitingHubLibrary,
    smiles_lookup: dict[str, str] | None = None,
) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title="waiting_hub_lnp")
    sheet.append(["Combo", "SMILES"])

    row_count = 0
    for row_count, row in enumerate(iter_waiting_hub_rows(library, smiles_lookup=smiles_lookup), start=1):
        sheet.append(list(row))
        if row_count % 2000 == 0:
            print(f"Wrote {row_count}/{TOTAL_ROWS} rows")

    if row_count != TOTAL_ROWS:
        raise ValueError(f"Output row count mismatch: got {row_count}, expected {TOTAL_ROWS}")

    workbook.save(output_path)


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    workbook_path = base_dir / INPUT_FILENAME
    output_path = base_dir / OUTPUT_FILENAME

    library = load_waiting_hub_library(workbook_path)

    if rdkit_available():
        smiles_lookup = None
        print("RDKit detected, generating SMILES from the reaction definition.")
    else:
        print("RDKit not available, falling back to existing full-library CSV lookups.")
        smiles_lookup = build_existing_smiles_lookup(library, base_dir.parent)

    export_waiting_hub_excel(output_path, library, smiles_lookup=smiles_lookup)
    print(output_path)


if __name__ == "__main__":
    main()
