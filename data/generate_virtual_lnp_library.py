from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from rdkit import Chem
from rdkit.Chem import rdchem


EXPECTED_COUNTS = {"A": 40, "B": 20, "C": 40, "D": 40}
STATUS_VALID = "Valid lipid structure"
TOTAL_ROWS_PER_FILE = 20 * 20 * 40 * 40


def canonicalize_smiles(smiles: str) -> str:
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError(f"Invalid SMILES: {smiles}")
    return Chem.MolToSmiles(mol)


def _find_isocyanide_atoms(mol: Chem.Mol) -> tuple[int, int]:
    match = mol.GetSubstructMatches(Chem.MolFromSmarts("[N+]#[C-]"))
    if not match:
        raise ValueError("Isocyanide group not found")
    return match[0]


def _find_carbonyl_atoms(mol: Chem.Mol) -> tuple[int, int]:
    for atom in mol.GetAtoms():
        if atom.GetAtomicNum() != 6:
            continue
        for bond in atom.GetBonds():
            other = bond.GetOtherAtom(atom)
            if (
                bond.GetBondType() == rdchem.BondType.DOUBLE
                and other.GetAtomicNum() == 8
            ):
                return atom.GetIdx(), other.GetIdx()
    raise ValueError("Carbonyl group not found")


def _find_primary_amine_n(mol: Chem.Mol) -> int:
    for atom in mol.GetAtoms():
        if atom.GetAtomicNum() == 7 and atom.GetTotalNumHs() >= 2:
            return atom.GetIdx()
    raise ValueError("Primary amine nitrogen not found")


def _find_carboxylic_acid_atoms(mol: Chem.Mol) -> tuple[int, int]:
    matches = mol.GetSubstructMatches(Chem.MolFromSmarts("[CX3](=O)[OX2H1]"))
    if not matches:
        raise ValueError("Carboxylic acid group not found")
    carbonyl_carbon, _, hydroxy_oxygen = matches[0]
    return carbonyl_carbon, hydroxy_oxygen


def _shift_after_removals(index: int, removed_indices: Sequence[int]) -> int:
    return index - sum(1 for removed in removed_indices if removed < index)


def generate_ugi_product_smiles(
    isocyanide_smiles: str,
    carbonyl_smiles: str,
    amine_smiles: str,
    acid_smiles: str,
) -> str:
    isocyanide = Chem.MolFromSmiles(isocyanide_smiles)
    carbonyl = Chem.MolFromSmiles(carbonyl_smiles)
    amine = Chem.MolFromSmiles(amine_smiles)
    acid = Chem.MolFromSmiles(acid_smiles)

    if None in (isocyanide, carbonyl, amine, acid):
        raise ValueError("Failed to parse one or more reactant SMILES")

    iso_n, iso_c = _find_isocyanide_atoms(isocyanide)
    carbonyl_c, carbonyl_o = _find_carbonyl_atoms(carbonyl)
    amine_n = _find_primary_amine_n(amine)
    acid_c, acid_oh_o = _find_carboxylic_acid_atoms(acid)

    combo = Chem.CombineMols(
        Chem.CombineMols(Chem.CombineMols(isocyanide, carbonyl), amine),
        acid,
    )
    offsets = [
        0,
        isocyanide.GetNumAtoms(),
        isocyanide.GetNumAtoms() + carbonyl.GetNumAtoms(),
        isocyanide.GetNumAtoms() + carbonyl.GetNumAtoms() + amine.GetNumAtoms(),
    ]

    iso_n += offsets[0]
    iso_c += offsets[0]
    carbonyl_c += offsets[1]
    carbonyl_o += offsets[1]
    amine_n += offsets[2]
    acid_c += offsets[3]
    acid_oh_o += offsets[3]

    rw = Chem.RWMol(combo)
    removed = sorted([carbonyl_o, acid_oh_o], reverse=True)
    for atom_idx in removed:
        rw.RemoveAtom(atom_idx)

    removed_low_to_high = sorted(removed)
    iso_n = _shift_after_removals(iso_n, removed_low_to_high)
    iso_c = _shift_after_removals(iso_c, removed_low_to_high)
    carbonyl_c = _shift_after_removals(carbonyl_c, removed_low_to_high)
    amine_n = _shift_after_removals(amine_n, removed_low_to_high)
    acid_c = _shift_after_removals(acid_c, removed_low_to_high)

    iso_bond = rw.GetBondBetweenAtoms(iso_n, iso_c)
    iso_bond.SetBondType(rdchem.BondType.SINGLE)

    iso_n_atom = rw.GetAtomWithIdx(iso_n)
    iso_n_atom.SetFormalCharge(0)
    iso_n_atom.SetNumExplicitHs(1)
    iso_n_atom.SetNoImplicit(True)
    rw.GetAtomWithIdx(iso_c).SetFormalCharge(0)

    oxygen_idx = rw.AddAtom(Chem.Atom(8))
    rw.AddBond(iso_c, oxygen_idx, rdchem.BondType.DOUBLE)
    rw.AddBond(iso_c, carbonyl_c, rdchem.BondType.SINGLE)
    rw.AddBond(amine_n, carbonyl_c, rdchem.BondType.SINGLE)
    rw.AddBond(amine_n, acid_c, rdchem.BondType.SINGLE)

    product = rw.GetMol()
    Chem.SanitizeMol(product)
    return canonicalize_smiles(Chem.MolToSmiles(product))


@dataclass(frozen=True)
class ComponentLibrary:
    a_block_1: list[str]
    a_block_2: list[str]
    b_values: list[str]
    c_values: list[str]
    d_values: list[str]


def _read_component_column(ws, column_index: int) -> list[str]:
    values: list[str] = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, column_index).value
        if value is not None:
            values.append(str(value).strip())
    return values


def load_component_library(workbook_path: Path) -> ComponentLibrary:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    headers = [worksheet.cell(1, col).value for col in range(2, 6)]
    if headers != ["A", "B", "C", "D"]:
        raise ValueError(f"Unexpected header row in {workbook_path}: {headers}")

    a_values = _read_component_column(worksheet, 2)
    b_values = _read_component_column(worksheet, 3)
    c_values = _read_component_column(worksheet, 4)
    d_values = _read_component_column(worksheet, 5)

    counts = {"A": len(a_values), "B": len(b_values), "C": len(c_values), "D": len(d_values)}
    if counts != EXPECTED_COUNTS:
        raise ValueError(f"Unexpected component counts: {counts}, expected {EXPECTED_COUNTS}")

    return ComponentLibrary(
        a_block_1=a_values[:20],
        a_block_2=a_values[20:40],
        b_values=b_values,
        c_values=c_values,
        d_values=d_values,
    )


def _combo_sort_order(prefix: str, size: int) -> list[int]:
    return sorted(range(1, size + 1), key=lambda idx: f"{prefix}{idx}")


def iter_virtual_rows(
    a_values: Sequence[str],
    b_values: Sequence[str],
    c_values: Sequence[str],
    d_values: Sequence[str],
) -> Iterable[tuple[str, str, str]]:
    a_order = _combo_sort_order("A", len(a_values))
    b_order = _combo_sort_order("B", len(b_values))
    c_order = _combo_sort_order("C", len(c_values))
    d_order = _combo_sort_order("D", len(d_values))

    for a_idx in a_order:
        a_smiles = a_values[a_idx - 1]
        for b_idx in b_order:
            b_smiles = b_values[b_idx - 1]
            for c_idx in c_order:
                c_smiles = c_values[c_idx - 1]
                for d_idx in d_order:
                    d_smiles = d_values[d_idx - 1]
                    combo = f"A{a_idx}B{b_idx}C{c_idx}D{d_idx}"
                    smiles = generate_ugi_product_smiles(
                        isocyanide_smiles=a_smiles,
                        carbonyl_smiles=b_smiles,
                        amine_smiles=c_smiles,
                        acid_smiles=d_smiles,
                    )
                    yield combo, smiles, STATUS_VALID


def write_library_csv(
    output_path: Path,
    a_values: Sequence[str],
    b_values: Sequence[str],
    c_values: Sequence[str],
    d_values: Sequence[str],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Combo", "SMILES", "Status"])

        row_count = 0
        for row_count, row in enumerate(
            iter_virtual_rows(a_values=a_values, b_values=b_values, c_values=c_values, d_values=d_values),
            start=1,
        ):
            writer.writerow(row)
            if row_count % 10000 == 0:
                print(f"[{output_path.name}] wrote {row_count}/{TOTAL_ROWS_PER_FILE} rows")

    if row_count != TOTAL_ROWS_PER_FILE:
        raise ValueError(
            f"{output_path.name} row count mismatch: got {row_count}, expected {TOTAL_ROWS_PER_FILE}"
        )


def main() -> None:
    data_dir = Path(__file__).resolve().parent
    workbook_path = data_dir / "virtual ingredient.xlsx"
    output_generated_1 = data_dir / "LNP_virtual_lipid_library_generated1.csv"
    output_generated_2 = data_dir / "LNP_virtual_lipid_library_generated2.csv"

    library = load_component_library(workbook_path)

    print("Generating generated2 from A1-A20 ...")
    write_library_csv(
        output_path=output_generated_2,
        a_values=library.a_block_1,
        b_values=library.b_values,
        c_values=library.c_values,
        d_values=library.d_values,
    )

    print("Generating generated1 from A21-A40 remapped to A1-A20 ...")
    write_library_csv(
        output_path=output_generated_1,
        a_values=library.a_block_2,
        b_values=library.b_values,
        c_values=library.c_values,
        d_values=library.d_values,
    )

    total_rows = TOTAL_ROWS_PER_FILE * 2
    print(f"Done. Generated {total_rows} rows across 2 files.")


if __name__ == "__main__":
    main()
