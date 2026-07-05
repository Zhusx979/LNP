from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from rdkit import Chem
from rdkit.Chem import rdchem


LOCAL_A_ROWS = range(24, 29)
LOCAL_C_ROWS = range(24, 32)
LOCAL_D_ROWS = range(24, 39)
LOCAL_B_ROWS = [24, 25]


def canonicalize_smiles(smiles: str) -> str:
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError(f"Invalid SMILES: {smiles}")
    return Chem.MolToSmiles(mol)


def _find_isocyanide_atoms(mol: Chem.Mol) -> Tuple[int, int]:
    match = mol.GetSubstructMatches(Chem.MolFromSmarts("[N+]#[C-]"))
    if not match:
        raise ValueError("Isocyanide group not found")
    return match[0]


def _find_carbonyl_atoms(mol: Chem.Mol) -> Tuple[int, int]:
    for atom in mol.GetAtoms():
        if atom.GetAtomicNum() != 6:
            continue
        for bond in atom.GetBonds():
            other = bond.GetOtherAtom(atom)
            if (
                bond.GetBondType() == rdchem.BondType.DOUBLE
                and other.GetAtomicNum() == 8
            ):
                return atom.GetIdx(), other.GetIdx()
    raise ValueError("Carbonyl group not found")


def _find_primary_amine_n(mol: Chem.Mol) -> int:
    for atom in mol.GetAtoms():
        if atom.GetAtomicNum() == 7 and atom.GetTotalNumHs() >= 2:
            return atom.GetIdx()
    raise ValueError("Primary amine nitrogen not found")


def _find_carboxylic_acid_atoms(mol: Chem.Mol) -> Tuple[int, int]:
    matches = mol.GetSubstructMatches(Chem.MolFromSmarts("[CX3](=O)[OX2H1]"))
    if not matches:
        raise ValueError("Carboxylic acid group not found")
    carbonyl_carbon, _, hydroxy_oxygen = matches[0]
    return carbonyl_carbon, hydroxy_oxygen


def _shift_after_removals(index: int, removed_indices: List[int]) -> int:
    return index - sum(1 for removed in removed_indices if removed < index)


def generate_ugi_product_smiles(
    isocyanide_smiles: str,
    carbonyl_smiles: str,
    amine_smiles: str,
    acid_smiles: str,
) -> str:
    isocyanide = Chem.MolFromSmiles(isocyanide_smiles)
    carbonyl = Chem.MolFromSmiles(carbonyl_smiles)
    amine = Chem.MolFromSmiles(amine_smiles)
    acid = Chem.MolFromSmiles(acid_smiles)

    if None in (isocyanide, carbonyl, amine, acid):
        raise ValueError("Failed to parse one or more reactant SMILES")

    iso_n, iso_c = _find_isocyanide_atoms(isocyanide)
    carbonyl_c, carbonyl_o = _find_carbonyl_atoms(carbonyl)
    amine_n = _find_primary_amine_n(amine)
    acid_c, acid_oh_o = _find_carboxylic_acid_atoms(acid)

    combo = Chem.CombineMols(
        Chem.CombineMols(Chem.CombineMols(isocyanide, carbonyl), amine),
        acid,
    )
    offsets = [
        0,
        isocyanide.GetNumAtoms(),
        isocyanide.GetNumAtoms() + carbonyl.GetNumAtoms(),
        isocyanide.GetNumAtoms() + carbonyl.GetNumAtoms() + amine.GetNumAtoms(),
    ]

    iso_n += offsets[0]
    iso_c += offsets[0]
    carbonyl_c += offsets[1]
    carbonyl_o += offsets[1]
    amine_n += offsets[2]
    acid_c += offsets[3]
    acid_oh_o += offsets[3]

    rw = Chem.RWMol(combo)
    removed = sorted([carbonyl_o, acid_oh_o], reverse=True)
    for atom_idx in removed:
        rw.RemoveAtom(atom_idx)

    removed_low_to_high = sorted(removed)
    iso_n = _shift_after_removals(iso_n, removed_low_to_high)
    iso_c = _shift_after_removals(iso_c, removed_low_to_high)
    carbonyl_c = _shift_after_removals(carbonyl_c, removed_low_to_high)
    amine_n = _shift_after_removals(amine_n, removed_low_to_high)
    acid_c = _shift_after_removals(acid_c, removed_low_to_high)

    iso_bond = rw.GetBondBetweenAtoms(iso_n, iso_c)
    iso_bond.SetBondType(rdchem.BondType.SINGLE)

    iso_n_atom = rw.GetAtomWithIdx(iso_n)
    iso_n_atom.SetFormalCharge(0)
    iso_n_atom.SetNumExplicitHs(1)
    iso_n_atom.SetNoImplicit(True)
    rw.GetAtomWithIdx(iso_c).SetFormalCharge(0)

    oxygen_idx = rw.AddAtom(Chem.Atom(8))
    rw.AddBond(iso_c, oxygen_idx, rdchem.BondType.DOUBLE)
    rw.AddBond(iso_c, carbonyl_c, rdchem.BondType.SINGLE)
    rw.AddBond(amine_n, carbonyl_c, rdchem.BondType.SINGLE)
    rw.AddBond(amine_n, acid_c, rdchem.BondType.SINGLE)

    product = rw.GetMol()
    Chem.SanitizeMol(product)
    return Chem.MolToSmiles(product)


def _read_real_data_sheet(real_data_path: Path):
    workbook = load_workbook(real_data_path, data_only=True)
    return workbook[workbook.sheetnames[0]]


def _extract_local_component_map(real_data_path: Path) -> Dict[str, Dict[int, str]]:
    ws = _read_real_data_sheet(real_data_path)
    return {
        "A": {i: ws.cell(row, 4).value for i, row in enumerate(LOCAL_A_ROWS, start=1)},
        "B": {i: ws.cell(row, 11).value for i, row in enumerate(LOCAL_B_ROWS, start=1)},
        "C": {i: ws.cell(row, 17).value for i, row in enumerate(LOCAL_C_ROWS, start=1)},
        "D": {i: ws.cell(row, 25).value for i, row in enumerate(LOCAL_D_ROWS, start=1)},
    }


def _iter_real_data_records(real_data_path: Path):
    ws = _read_real_data_sheet(real_data_path)

    for d_idx, row in enumerate(range(4, 19), start=1):
        d_label = ws.cell(row, 1).value
        if d_label != f"D{d_idx}":
            raise ValueError(f"Unexpected D label at row {row}: {d_label}")

        for b_idx, start_col in enumerate([3, 43], start=1):
            for a_idx in range(1, 6):
                for c_idx in range(1, 9):
                    col = start_col + (a_idx - 1) * 8 + (c_idx - 1)
                    combo = f"A{a_idx}B{b_idx}C{c_idx}D{d_idx}"
                    yield {
                        "combo": combo,
                        "a_idx": a_idx,
                        "b_idx": b_idx,
                        "c_idx": c_idx,
                        "d_idx": d_idx,
                        "target": ws.cell(row, col).value,
                    }


def build_real_data_dataframe(real_data_path: Path | str) -> pd.DataFrame:
    real_data_path = Path(real_data_path)
    component_map = _extract_local_component_map(real_data_path)

    records = []
    for record in _iter_real_data_records(real_data_path):
        smiles = generate_ugi_product_smiles(
            isocyanide_smiles=component_map["A"][record["a_idx"]],
            carbonyl_smiles=component_map["B"][record["b_idx"]],
            amine_smiles=component_map["C"][record["c_idx"]],
            acid_smiles=component_map["D"][record["d_idx"]],
        )
        records.append(
            {
                "combo": record["combo"],
                "smiles": smiles,
                "target": record["target"],
            }
        )

    return pd.DataFrame(records, columns=["combo", "smiles", "target"])


@dataclass
class RealDataExporter:
    real_data_path: Path

    def build_dataframe(self) -> pd.DataFrame:
        return build_real_data_dataframe(self.real_data_path)

    def export_excel(self, output_path: Path | str) -> Path:
        output_path = Path(output_path)
        df = self.build_dataframe()[["smiles", "target", "combo"]]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(output_path, index=False)
        return output_path
